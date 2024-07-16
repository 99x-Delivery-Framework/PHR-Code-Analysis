import openpyxl
import csv
from collections import defaultdict
import os

# Define the categories and the keywords
keyword_categories = {
    "Cyclomatic Complexity": ["cognitive", "cyclomatic"],
    "Code Documentation": [],
    "Code Duplication": ["duplicate", "duplicates"],
    "Code Churn": [],
    "Code Coverage": [],
    "Code Security": [],
}

# Define severity mapping
severity_mapping = {
    "Cyclomatic Complexity": {
        "CRITICAL": "High Risk",
        "BLOCKER": "High Risk",
        "MAJOR": "Complex",
        "DEFAULT": "Moderate"
    },
    "Code Security": {
        "CRITICAL": "Unacceptable",
        "BLOCKER": "Unacceptable",
        "MAJOR": "Manage",
        "DEFAULT": "Acceptable"
    },
    "Code Duplication": {
        "CRITICAL": "High",
        "BLOCKER": "High",
        "MAJOR": "Medium",
        "DEFAULT": "Low"
    },
    "Code Bug Issues": {
        "CRITICAL": "High",
        "BLOCKER": "High",
        "MAJOR": "Medium",
        "DEFAULT": "Low"
    }
}

# Initialize a dictionary to count occurrences, with nested dictionaries for severities
category_counts = defaultdict(lambda: defaultdict(int))

# Function to read the sheet from an Excel file
def read_sheet_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in {file_path}.")
        return

    sheet = workbook[sheet_name]

    # Define the columns
    message_column = 'message'
    type_column = 'type'
    severity_column = 'severity'

    # Find column indices
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    
    try:
        message_idx = header.index(message_column)
        type_idx = header.index(type_column)
        severity_idx = header.index(severity_column)
    except ValueError as e:
        print(f"Column not found: {e}")
        return

    # Iterate through the rows and categorize issues
    for row in sheet.iter_rows(min_row=2, values_only=True):
        message = row[message_idx]
        issue_type = row[type_idx]
        severity = row[severity_idx]

        # Check if the type is BUG
        if issue_type == "BUG":
            mapped_severity = severity_mapping["Code Bug Issues"].get(severity, severity_mapping["Code Bug Issues"]["DEFAULT"])
            category_counts["Code Bug Issues"][mapped_severity] += 1
        else:
            # Check for keywords in the message
            categorized = False
            for category, keywords in keyword_categories.items():
                if any(keyword in message.lower() for keyword in keywords):
                    mapped_severity = severity_mapping[category].get(severity, severity_mapping[category]["DEFAULT"])
                    category_counts[category][mapped_severity] += 1
                    categorized = True
                    break

            # If no keyword match, add to a general "Other" category
            if not categorized:
                category_counts["Other"][severity] += 1

# Function to read the 'in' sheet from a CSV file
def read_all_sheet_csv(file_path):
    with open(file_path, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        
        # Check if required columns exist
        required_columns = ['message', 'severity']
        if not all(col in reader.fieldnames for col in required_columns):
            print(f"One or more required columns {required_columns} not found in {file_path}.")
            return
        
        # Iterate through the rows and categorize issues
        for row in reader:
            message = row['message']
            severity = row['severity']

            # Check for keywords in the message
            categorized = False
            for category, keywords in keyword_categories.items():
                if any(keyword in message.lower() for keyword in keywords):
                    mapped_severity = severity_mapping[category].get(severity, severity_mapping[category]["DEFAULT"])
                    category_counts[category][mapped_severity] += 1
                    categorized = True
                    break

                # If no keyword match, add to a general "Uncategorized" category (optional)
                if not categorized:
                    category_counts["Other"][severity] += 1

# Function to read the 'Security Hotspots' sheet from the Excel file
def read_security_hotspot(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in {file_path}.")
        return

    sheet = workbook[sheet_name]

    # Define the columns
    severity_column = 'Severity'

    # Find column indices
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    
    try:
        severity_idx = header.index(severity_column)
    except ValueError as e:
        print(f"Column not found: {e}")
        return

    # Iterate through the rows and categorize issues
    for row in sheet.iter_rows(min_row=2, values_only=True):
        severity = row[severity_idx]
        mapped_severity = severity_mapping["Code Security"].get(severity, severity_mapping["Code Security"]["DEFAULT"])
        category_counts["Code Security"][mapped_severity] += 1

# Function to save results to Excel
def save_to_excel(output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"

    # Write headers
    sheet.append(["Category", "Severity", "Count"])

    # Define sort orders
    sort_orders = {
        "Cyclomatic Complexity": ["High Risk", "Complex", "Moderate"],
        "Code Security": ["Unacceptable", "Manage", "Acceptable"],
        "Code Duplication": ["High", "Medium", "Low"],
        "Code Bug Issues": ["High", "Medium", "Low"]
    }

    # Write data
    for category, severities in category_counts.items():
        sorted_severities = sorted(severities.items(), key=lambda x: sort_orders[category].index(x[0]) if category in sort_orders else x[0])
        for severity, count in sorted_severities:
            sheet.append([category, severity, count])

    # Save the workbook
    workbook.save(output_file)

# Function to generate output file name
def get_output_file_name(input_file):
    base, ext = os.path.splitext(input_file)
    return f"{base}-summary{ext}"

# Function to process a single file
def process_file(input_file):
    # Clear category counts for each new file
    global category_counts
    category_counts = defaultdict(lambda: defaultdict(int))

    output_file = get_output_file_name(input_file)
    all_sheet_name = 'All'
    security_hotspot_sheet_name = 'Security Hotspots'

    # Determine file type and process accordingly
    file_extension = os.path.splitext(input_file)[1]

    if file_extension == '.xlsx':
        # Read and categorize issues from the 'All' sheet in Excel
        read_sheet_excel(input_file, all_sheet_name)

        # Read and categorize issues from the 'Security Hotspot' sheet in Excel
        read_security_hotspot(input_file, security_hotspot_sheet_name)
    elif file_extension == '.csv':
        # Read and categorize issues from the 'in' sheet in CSV
        read_all_sheet_csv(input_file)
    else:
        print(f"Unsupported file type: {file_extension}")

    save_to_excel(output_file)

input_folder = 'input'  # Folder containing the input files

# Iterate over all files in the input folder
for filename in os.listdir(input_folder):
    if filename.endswith('.xlsx') or filename.endswith('.csv'):
        input_file = os.path.join(input_folder, filename)
        print(f"Processing file: {input_file}")
        process_file(input_file)